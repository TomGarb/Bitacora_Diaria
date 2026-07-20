from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
import openpyxl
import io
import uuid

from app.database import get_db
# Agregamos AppUser a las importaciones
from app.models.shift import Client, DataCenterEntity, AppUser
from app.auth import get_current_user

router = APIRouter(prefix="/master-data", tags=["Master Data Management"])
templates = Jinja2Templates(directory="app/templates")

@router.get("/")
async def render_master_data(request: Request, user_info: dict = Depends(get_current_user)):
    nombre_real = user_info.get("preferred_username", "Operador")
    rol_local = user_info.get("local_role", "Operador")  # <-- Capturamos el rol de la BD local
    
    return templates.TemplateResponse("master_data.html", {
        "request": request, 
        "user_name": nombre_real,
        "user_role": rol_local  # <-- Pasamos el rol al HTML
    })

# ==========================================
# APIS DE CONSULTA (LISTAR)
# ==========================================
@router.get("/api/clients")
def get_clients(db: Session = Depends(get_db)):
    return db.query(Client).order_by(Client.name).all()

@router.get("/api/datacenters")
def get_datacenters(db: Session = Depends(get_db)):
    return db.query(DataCenterEntity).order_by(DataCenterEntity.name).all()

# ==========================================
# APIS DE CARGA INDIVIDUAL
# ==========================================
@router.post("/api/client")
def add_client(data: dict, db: Session = Depends(get_db)):
    if db.query(Client).filter(Client.code == data["code"]).first():
        raise HTTPException(status_code=400, detail="El código de cliente ya existe.")
    client = Client(name=data["name"], code=data["code"])
    db.add(client)
    db.commit()
    return {"message": "Cliente guardado"}

@router.post("/api/datacenter")
def add_datacenter(data: dict, db: Session = Depends(get_db)):
    if db.query(DataCenterEntity).filter(DataCenterEntity.name == data["name"]).first():
        raise HTTPException(status_code=400, detail="El DataCenter ya existe.")
    dc = DataCenterEntity(name=data["name"])
    db.add(dc)
    db.commit()
    return {"message": "DataCenter guardado"}

# ==========================================
# APIS DE MODIFICACIÓN (PUT)
# ==========================================
@router.put("/api/client/{client_id}")
def update_client(client_id: str, data: dict, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    duplicate = db.query(Client).filter(Client.code == data["code"], Client.id != client_id).first()
    if duplicate:
        raise HTTPException(status_code=400, detail="Ese código ya está asignado a otro cliente")

    client.name = data["name"]
    client.code = data["code"]
    db.commit()
    return {"message": "Cliente modificado correctamente"}

@router.put("/api/datacenter/{dc_id}")
def update_datacenter(dc_id: str, data: dict, db: Session = Depends(get_db)):
    dc = db.query(DataCenterEntity).filter(DataCenterEntity.id == dc_id).first()
    if not dc:
        raise HTTPException(status_code=404, detail="DataCenter no encontrado")

    dc.name = data["name"]
    db.commit()
    return {"message": "DataCenter modificado correctamente"}

# ==========================================
# APIS DE ELIMINACIÓN (DELETE)
# ==========================================
@router.delete("/api/client/{client_id}")
def delete_client(client_id: str, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    db.delete(client)
    db.commit()
    return {"message": "Cliente eliminado con éxito"}

@router.delete("/api/datacenter/{dc_id}")
def delete_datacenter(dc_id: str, db: Session = Depends(get_db)):
    dc = db.query(DataCenterEntity).filter(DataCenterEntity.id == dc_id).first()
    if not dc:
        raise HTTPException(status_code=404, detail="DataCenter no encontrado")
    db.delete(dc)
    db.commit()
    return {"message": "DataCenter eliminado con éxito"}

# ==========================================
# EXCEL BULK UPLOAD
# ==========================================
@router.post("/api/bulk/clients")
async def bulk_upload_clients(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0] and row[1]:
            if not db.query(Client).filter(Client.code == str(row[1])).first():
                db.add(Client(name=str(row[0]), code=str(row[1])))
                count += 1
    db.commit()
    return {"message": f"Se registraron {count} nuevos clientes correctamente."}

@router.post("/api/bulk/datacenters")
async def bulk_upload_datacenters(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet = wb.active
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            if not db.query(DataCenterEntity).filter(DataCenterEntity.name == str(row[0])).first():
                db.add(DataCenterEntity(name=str(row[0])))
                count += 1
    db.commit()
    return {"message": f"Se registraron {count} nuevos DataCenters correctamente."}

# ==========================================
# GENERADORES DE PLANTILLAS EXCEL
# ==========================================
@router.get("/api/template/clients")
def download_clients_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["Nombre", "Codigo"])
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {'Content-Disposition': 'attachment; filename="plantilla_clientes.xlsx"'}
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@router.get("/api/template/datacenters")
def download_datacenters_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataCenters"
    ws.append(["Nombre"])
    ws.column_dimensions['A'].width = 50
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {'Content-Disposition': 'attachment; filename="plantilla_datacenters.xlsx"'}
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

# ==========================================
# API DE USUARIOS PROTEGIDA
# ==========================================
@router.get("/api/users")
def get_local_users(db: Session = Depends(get_db), user_info: dict = Depends(get_current_user)):
    # VALIDACIÓN: Si no es Administrador, rebota la petición inmediatamente
    if user_info.get("local_role") != "Administrador":
        raise HTTPException(status_code=403, detail="Acceso denegado: Se requieren permisos de Administrador.")
        
    usuarios = db.query(AppUser).order_by(AppUser.username.asc()).all()
    return [{
        "id": u.id,
        "username": u.username,
        "firstName": u.username,
        "lastName": "",
        "email": "Sincronizado vía SSO",
        "enabled": True,
        "role": u.role
    } for u in usuarios]

@router.post("/api/users")
def save_app_user(
    id: str = Form(""),
    username: str = Form(...),
    role: str = Form(...),
    db: Session = Depends(get_db),
    user_info: dict = Depends(get_current_user)
):
    # VALIDACIÓN: Evita que un operador intente enviar un POST mediante herramientas externas (Postman/Curl)
    if user_info.get("local_role") != "Administrador":
        raise HTTPException(status_code=403, detail="Acceso denegado: No tienes permisos para modificar usuarios.")
        
    if id:
        user = db.query(AppUser).filter(AppUser.id == id).first()
        if user:
            user.username = username
            user.role = role
    else:
        existing = db.query(AppUser).filter(AppUser.username == username).first()
        if existing:
            raise HTTPException(status_code=400, detail="El usuario ya existe.")
        new_user = AppUser(id=str(uuid.uuid4()), username=username, role=role)
        db.add(new_user)
    
    db.commit()
    return {"message": "Usuario guardado exitosamente."}

@router.delete("/api/users/{user_id}")
def delete_app_user(user_id: str, db: Session = Depends(get_db), user_info: dict = Depends(get_current_user)):
    # VALIDACIÓN: Protege la eliminación
    if user_info.get("local_role") != "Administrador":
        raise HTTPException(status_code=403, detail="Acceso denegado: No tienes permisos para eliminar usuarios.")
        
    user = db.query(AppUser).filter(AppUser.id == user_id).first()
    if user:
        db.delete(user)
        db.commit()
    return {"message": "Usuario eliminado."}