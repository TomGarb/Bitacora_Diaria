from fastapi import APIRouter, Request, Depends, UploadFile, File, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
import io
import uuid
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

from app.database import get_db
from app.models.shift import ShiftMetrics

router = APIRouter(prefix="/metrics", tags=["Métricas e Indicadores"])
templates = Jinja2Templates(directory="app/templates")

WEEKDAYS_ES = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"}
MONTHS_ES = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}

@router.get("/")
def render_metrics_dashboard(request: Request):
    return templates.TemplateResponse("metrics.html", {"request": request})

@router.get("/api/data")
def get_metrics_data(db: Session = Depends(get_db)):
    # Retorna los últimos 30 registros para graficar
    return db.query(ShiftMetrics).order_by(ShiftMetrics.created_at.desc()).limit(30).all()

@router.post("/api/save")
def save_shift_metrics(request: Request, db: Session = Depends(get_db), data: dict = None):
    # Endpoint para capturar las métricas enviadas desde el modal de cierre de turno
    import asyncio
    # Procesar json directamente
    async def get_body():
        return await request.json()
    
    body = asyncio.run(get_body())
    m = body.get("metrics", {})
    
    new_metrics = ShiftMetrics(
        id=str(uuid.uuid4()),
        operator_name=body.get("operator_name", "Operador Desconocido"),
        shift=body.get("shift", "No especificado"),
        created_at=datetime.utcnow(),
        # Mapeo masivo de valores seguros
        sh_aws_c=int(m.get("sh_aws", {}).get("c") or 0), sh_aws_t=int(m.get("sh_aws", {}).get("t") or 0),
        sh_c=int(m.get("sh", {}).get("c") or 0), sh_t=int(m.get("sh", {}).get("t") or 0),
        rh_c=int(m.get("rh", {}).get("c") or 0), rh_t=int(m.get("rh", {}).get("t") or 0),
        inout_c=int(m.get("inout", {}).get("c") or 0), inout_t=int(m.get("inout", {}).get("t") or 0),
        inc_c=int(m.get("inc", {}).get("c") or 0), inc_t=int(m.get("inc", {}).get("t") or 0),
        inc_aws_c=int(m.get("inc_aws", {}).get("c") or 0), inc_aws_t=int(m.get("inc_aws", {}).get("t") or 0),
        calls_c=int(m.get("calls", {}).get("c") or 0), calls_t=int(m.get("calls", {}).get("t") or 0),
        mails_c=int(m.get("mails", {}).get("c") or 0), mails_t=int(m.get("mails", {}).get("t") or 0),
        td_c=int(m.get("td", {}).get("c") or 0), td_t=int(m.get("td", {}).get("t") or 0),
        acc_c=int(m.get("acc", {}).get("c") or 0), acc_t=int(m.get("acc", {}).get("t") or 0),
        trad_c=int(m.get("trad", {}).get("c") or 0), trad_t=int(m.get("trad", {}).get("t") or 0),
        pta_c=int(m.get("pta", {}).get("c") or 0), pta_t=int(m.get("pta", {}).get("t") or 0),
        mop_c=int(m.get("mop", {}).get("c") or 0), mop_t=int(m.get("mop", {}).get("t") or 0),
        abm_c=int(m.get("abm", {}).get("c") or 0), abm_t=int(m.get("abm", {}).get("t") or 0),
        eti_c=int(m.get("eti", {}).get("c") or 0), eti_t=int(m.get("eti", {}).get("t") or 0),
        chi_c=int(m.get("chi", {}).get("c") or 0), chi_t=int(m.get("chi", {}).get("t") or 0),
        mia_c=int(m.get("mia", {}).get("c") or 0), mia_t=int(m.get("mia", {}).get("t") or 0),
        te_c=int(m.get("te", {}).get("c") or 0), te_t=int(m.get("te", {}).get("t") or 0),
    )
    db.add(new_metrics)
    db.commit()
    return {"status": "ok"}

@router.get("/export/excel", tags=["Metricas"])
def exportar_reporte_global_excel(db: Session = Depends(get_db)):
    try:
        # 1. Extraer los datos reales de PostgreSQL
        registros = db.query(ShiftMetrics).order_by(ShiftMetrics.created_at.asc()).all()
        
        wb = openpyxl.Workbook()
        
        # Estilos visuales profesionales
        font_family = "Segoe UI"
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        sub_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        accent_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        
        title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        bold_font = Font(name=font_family, size=11, bold=True, color="0F172A")
        regular_font = Font(name=font_family, size=10, color="334155")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0")
        )

        # --- PESTAÑA 1: DASHBOARD INTERACTIVO ---
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        ws_dash.views.sheetView[0].showGridLines = False
        
        ws_dash.merge_cells("A1:M1")
        t_cell = ws_dash["A1"]
        t_cell.value = "📊 PANEL DE CONTROL ANALÍTICO - BITÁCORA DIARIA"
        t_cell.fill = header_fill
        t_cell.font = title_font
        t_cell.alignment = center_align
        ws_dash.row_dimensions[1].height = 40
        
        ws_dash.cell(row=3, column=2, value="Seleccionar Turno:").font = bold_font
        ws_dash.cell(row=3, column=2).alignment = right_align
        filter_cell = ws_dash.cell(row=3, column=3, value="Todos")
        filter_cell.font = Font(name=font_family, size=11, bold=True, color="1E3A8A")
        filter_cell.fill = accent_fill
        filter_cell.alignment = center_align
        filter_cell.border = thin_border
        
        dv = DataValidation(type="list", formula1='"Todos,Mañana,Tarde,Noche,Central"', allow_blank=False)
        ws_dash.add_data_validation(dv)
        dv.add(filter_cell)

        # --- PESTAÑA MOTOR DE DATOS (Oculta) ---
        ws_flat = wb.create_sheet(title="Datos_Procesados")
        headers_flat = [
            "Fecha", "Día", "Turno", "Feriado", "Total Tareas", "Tiempo Total", "A100 Tareas", "Otros Tareas",
            "sh_aws_t", "sh_t", "rh_t", "inout_t", "inc_t", "inc_aws_t", "calls_t", "mails_t", "td_t", "acc_t", "trad_t", "pta_t", "mop_t", "abm_t", "eti_t", "chi_t", "mia_t", "te_t",
            "sh_aws_c", "inc_aws_c", "chi_c", "mop_c"
        ]
        for c_idx, h in enumerate(headers_flat, 1):
            ws_flat.cell(row=1, column=c_idx, value=h).fill = header_fill

        # --- PESTAÑAS TRADICIONALES DE TURNOS ---
        hojas_turnos = {
            "TM": wb.create_sheet(title="TM"), "TT": wb.create_sheet(title="TT"),
            "TN": wb.create_sheet(title="TN"), "TC": wb.create_sheet(title="TC")
        }
        
        headers_tradicionales = [
            "Día", "Feriado", "Fecha / Planilla", "Fecha Efectiva", "Mes", "SH AWS", "SH", "RH", "IN/OUT", "INC", 
            "Alarmas/INC AWS", "Calls", "Mails", "Tareas Diarias", "Solicitud de acceso/Visitas/Escoltar de clientes", 
            "Traducción entre sectores", "Asistencia de puerta (llaves/elementos,etc)", "Seguimiento de mantenimiento/Solicitud de MOP", 
            "ABM de Equipos CMDB/ASSET", "ABM de Etiquetas", "Llamadas/Mails AWS Chile", "Llamadas/Casos MIA", "Tareas Adicionales", 
            "Break", "Total Tiempo empleado", "Total Ocioso"
        ]
        
        for sheet in hojas_turnos.values():
            sheet.views.sheetView[0].showGridLines = True
            for c_idx, h in enumerate(headers_tradicionales, 1):
                cell = sheet.cell(row=1, column=c_idx, value=h)
                cell.fill = sub_header_fill
                cell.font = header_font
                cell.alignment = center_align
            sheet.row_dimensions[1].height = 28

        f_idx = 2
        lineas_por_hoja = {"TM": 2, "TT": 2, "TN": 2, "TC": 2}
        mapa_nombres_hojas = {"Mañana": "TM", "Tarde": "TT", "Noche": "TN", "Central": "TC"}
        dias_semana = ["Domingo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

        for r in registros:
            dt = r.created_at if r.created_at else datetime.utcnow()
            fecha_str = dt.strftime("%d/%m/%Y")
            dia_str = dias_semana[int(dt.strftime("%w"))]
            mes_int = int(dt.strftime("%m"))
            
            # Evaluar Checkbox de Feriado (☑ si es 'on' o True, ☐ si no)
            val_feriado_raw = getattr(r, "feriado", "")
            is_feriado_box = "☑" if val_feriado_raw == "on" or val_feriado_raw is True else "☐"
            
            # --- Escritura en Hoja Plana de Soporte ---
            ws_flat.cell(row=f_idx, column=1, value=dt.strftime("%Y-%m-%d"))
            ws_flat.cell(row=f_idx, column=2, value=dia_str)
            ws_flat.cell(row=f_idx, column=3, value=r.shift)
            ws_flat.cell(row=f_idx, column=4, value=is_feriado_box)
            
            # Fórmulas con SUM para evitar errores #¡VALOR! con las celdas vacías
            ws_flat.cell(row=f_idx, column=5, value=f"=SUM(E{f_idx},G{f_idx},I{f_idx},K{f_idx},M{f_idx},O{f_idx},Q{f_idx},S{f_idx},U{f_idx},W{f_idx},Y{f_idx},AA{f_idx},AC{f_idx},AE{f_idx},AG{f_idx},AI{f_idx},AK{f_idx},AM{f_idx})")
            ws_flat.cell(row=f_idx, column=6, value=f"=SUM(F{f_idx},H{f_idx},J{f_idx},L{f_idx},N{f_idx},P{f_idx},R{f_idx},T{f_idx},V{f_idx},X{f_idx},Z{f_idx},AB{f_idx},AD{f_idx},AF{f_idx},AH{f_idx},AJ{f_idx},AL{f_idx},AN{f_idx})")
            ws_flat.cell(row=f_idx, column=7, value=f"=SUM(AA{f_idx},AB{f_idx},AC{f_idx},AD{f_idx})")
            ws_flat.cell(row=f_idx, column=8, value=f"=E{f_idx}-G{f_idx}")
            
            tiempos_list = [r.sh_aws_t, r.sh_t, r.rh_t, r.inout_t, r.inc_t, r.inc_aws_t, r.calls_t, r.mails_t, r.td_t, r.acc_t, r.trad_t, r.pta_t, r.mop_t, r.abm_t, r.eti_t, r.chi_t, r.mia_t, r.te_t]
            for pos, val in enumerate(tiempos_list):
                ws_flat.cell(row=f_idx, column=9 + pos, value=val if val != 0 else "")
                
            ws_flat.cell(row=f_idx, column=27, value=r.sh_aws_c if r.sh_aws_c != 0 else "")
            ws_flat.cell(row=f_idx, column=28, value=r.inc_aws_c if r.inc_aws_c != 0 else "")
            ws_flat.cell(row=f_idx, column=29, value=r.chi_c if r.chi_c != 0 else "")
            ws_flat.cell(row=f_idx, column=30, value=r.mop_c if r.mop_c != 0 else "")
            f_idx += 1

            # --- Escritura en Hojas Tradicionales ---
            h_nombre = mapa_nombres_hojas.get(r.shift)
            if h_nombre:
                ws_t = hojas_turnos[h_nombre]
                row_c = lineas_por_hoja[h_nombre]
                row_t = row_c + 1
                
                # Columnas comunes fijas
                ws_t.cell(row=row_c, column=1, value=dia_str)
                ws_t.cell(row=row_c, column=2, value=is_feriado_box) # Casilla ☑ o ☐
                ws_t.cell(row=row_c, column=3, value=fecha_str)
                ws_t.cell(row=row_c, column=4, value=fecha_str)
                ws_t.cell(row=row_c, column=5, value=mes_int)
                
                ws_t.cell(row=row_t, column=1, value=dia_str)
                ws_t.cell(row=row_t, column=2, value=is_feriado_box) # Casilla ☑ o ☐
                ws_t.cell(row=row_t, column=3, value=f"TD{dt.strftime('%y%m%d')}")
                ws_t.cell(row=row_t, column=4, value=fecha_str)
                ws_t.cell(row=row_t, column=5, value=mes_int)
                
                mapeo_metricas = [
                    (r.sh_aws_c, r.sh_aws_t), (r.sh_c, r.sh_t), (r.rh_c, r.rh_t), (r.inout_c, r.inout_t),
                    (r.inc_c, r.inc_t), (r.inc_aws_c, r.inc_aws_t), (r.calls_c, r.calls_t), (r.mails_c, r.mails_t),
                    (r.td_c, r.td_t), (r.acc_c, r.acc_t), (r.trad_c, r.trad_t), (r.pta_c, r.pta_t),
                    (r.mop_c, r.mop_t), (r.abm_c, r.abm_t), (r.eti_c, r.eti_t), (r.chi_c, r.chi_t),
                    (r.mia_c, r.mia_t), (r.te_c, r.te_t)
                ]
                
                for idx_m, (c_val, t_val) in enumerate(mapeo_metricas):
                    col_dest = 6 + idx_m
                    
                    # Si el valor es 0 o None, se escribe cadena vacía "" (Celda Limpia)
                    val_c = c_val if (c_val is not None and c_val != 0) else ""
                    cell_c = ws_t.cell(row=row_c, column=col_dest, value=val_c)
                    if val_c != "": cell_c.number_format = "#,##0"
                        
                    val_t = t_val if (t_val is not None and t_val != 0) else ""
                    cell_t = ws_t.cell(row=row_t, column=col_dest, value=val_t)
                    if val_t != "": cell_t.number_format = "#,##0"
                
                # Configuración de los Breaks permanentes
                ws_t.cell(row=row_c, column=24, value=1).number_format = "#,##0"
                ws_t.cell(row=row_t, column=24, value=60).number_format = "#,##0"
                
                # Totales con función SUM para tolerar celdas en blanco
                ws_t.cell(row=row_t, column=25, value=f"=SUM(F{row_t}:X{row_t})")
                ws_t.cell(row=row_t, column=26, value=f"=420-Y{row_t}")
                
                for r_num in [row_c, row_t]:
                    for c_num in range(1, 27):
                        cell = ws_t.cell(row=r_num, column=c_num)
                        cell.font = regular_font
                        cell.border = thin_border
                        if c_num >= 2: cell.alignment = center_align
                
                lineas_por_hoja[h_nombre] += 2

        for sheet in hojas_turnos.values():
            for col in sheet.columns:
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = 14

        # --- PESTAÑA DE CALCULOS (Alimentación de Gráficos) ---
        ws_calc = wb.create_sheet(title="Resumen_Calculos")
        ws_calc.cell(row=1, column=1, value="Fecha"); ws_calc.cell(row=1, column=2, value="Real Tareas")
        ws_calc.cell(row=1, column=3, value="Meta"); ws_calc.cell(row=1, column=4, value="Tiempo Total")
        ws_calc.cell(row=1, column=5, value="Tiempo Libre")
        
        fechas_unicas = sorted(list(set([ws_flat.cell(row=x, column=1).value for x in range(2, f_idx)]))) if f_idx > 2 else []
        for i, f_val in enumerate(fechas_unicas, 2):
            ws_calc.cell(row=i, column=1, value=f_val)
            ws_calc.cell(row=i, column=2, value=f'=IF(Dashboard!$C$3="Todos", SUMIF(Datos_Procesados!$A:$A, A{i}, Datos_Procesados!$E:$E), SUMIFS(Datos_Procesados!$E:$E, Datos_Procesados!$A:$A, A{i}, Datos_Procesados!$C:$C, Dashboard!$C$3))')
            ws_calc.cell(row=i, column=3, value=40)
            ws_calc.cell(row=i, column=4, value=f'=IF(Dashboard!$C$3="Todos", SUMIF(Datos_Procesados!$A:$A, A{i}, Datos_Procesados!$F:$F), SUMIFS(Datos_Procesados!$F:$F, Datos_Procesados!$A:$A, A{i}, Datos_Procesados!$C:$C, Dashboard!$C$3))')
            ws_calc.cell(row=i, column=5, value=f'=MAX(0, IF(Dashboard!$C$3="Todos", 1680, 420) - D{i})')

        # Distribución de Tiempos Macro
        ws_calc.cell(row=20, column=1, value="Categoría"); ws_calc.cell(row=20, column=2, value="Minutos")
        macros_labels = ["Hardware/RH", "Comunicaciones", "Físico DC", "Incidentes", "Gestión/Otros"]
        formulas_macros = [
            '=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!$I:$K), SUMIFS(Datos_Procesados!$I:$I, Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$J:$J, Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$K:$K, Datos_Procesados!$C:$C, Dashboard!$C$3))',
            '=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!$O:$P), SUMIFS(Datos_Procesados!$O:$O, Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$P:$P, Datos_Procesados!$C:$C, Dashboard!$C$3))',
            '=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!$L:$N), SUMIFS(Datos_Procesados!$L:$L, Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$M:$M, Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$N:$N, Datos_Procesados!$C:$C, Dashboard!$C$3))',
            '=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!$Q:$R), SUMIFS(Datos_Procesados!$Q:$Q, Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$R:$R, Datos_Procesados!$C:$C, Dashboard!$C$3))',
            '=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!$S:$Z), SUMIFS(Datos_Procesados!$S:$S, Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$T:$T, Datos_Procesados!$C:$C, Dashboard!$C$3))'
        ]
        for idx, label in enumerate(macros_labels):
            ws_calc.cell(row=21+idx, column=1, value=label)
            ws_calc.cell(row=21+idx, column=2, value=formulas_macros[idx])

        # A100 vs Otros
        ws_calc.cell(row=28, column=1, value="Grupo"); ws_calc.cell(row=28, column=2, value="Volumen")
        ws_calc.cell(row=29, column=1, value="Proyecto A100"); ws_calc.cell(row=29, column=2, value='=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!$G:$G), SUMIFS(Datos_Procesados!$G:$G, Datos_Procesados!$C:$C, Dashboard!$C$3))')
        ws_calc.cell(row=30, column=1, value="Resto Operativa"); ws_calc.cell(row=30, column=2, value='=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!$H:$H), SUMIFS(Datos_Procesados!$H:$H, Datos_Procesados!$C:$C, Dashboard!$C$3))')

        # Desglose A100
        ws_calc.cell(row=33, column=1, value="Tarea"); ws_calc.cell(row=33, column=2, value="Vol")
        dg_labels = ["SH AWS", "Incidentes AWS", "Llamadas Chile", "Mantenimientos/MOPs"]
        dg_cols = ["AA", "AB", "AC", "AD"]
        for idx, label in enumerate(dg_labels):
            ws_calc.cell(row=34+idx, column=1, value=label)
            ws_calc.cell(row=34+idx, column=2, value=f'=IF(Dashboard!$C$3="Todos", SUM(Datos_Procesados!${dg_cols[idx]}:${dg_cols[idx]}), SUMIFS(Datos_Procesados!${dg_cols[idx]}:${dg_cols[idx]}, Datos_Procesados!$C:$C, Dashboard!$C$3))')

        # Ritmo Semanal
        ws_calc.cell(row=40, column=1, value="Día"); ws_calc.cell(row=40, column=2, value="Volumen Semanal")
        for idx, d_name in enumerate(["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]):
            ws_calc.cell(row=41+idx, column=1, value=d_name)
            ws_calc.cell(row=41+idx, column=2, value=f'=IF(Dashboard!$C$3="Todos", SUMIF(Datos_Procesados!$B:$B, A{41+idx}, Datos_Procesados!$E:$E), SUMIFS(Datos_Procesados!$E:$E, Datos_Procesados!$B:$B, A{41+idx}, Datos_Procesados!$C:$C, Dashboard!$C$3))')

        # Fin de semana
        ws_calc.cell(row=50, column=1, value="Tipo"); ws_calc.cell(row=50, column=2, value="Promedio")
        ws_calc.cell(row=51, column=1, value="Promedio Lun-Vie")
        ws_calc.cell(row=51, column=2, value='=IF(Dashboard!$C$3="Todos", AVERAGEIFS(Datos_Procesados!$E:$E, Datos_Procesados!$B:$B, "<>Sábado", Datos_Procesados!$B:$B, "<>Domingo"), AVERAGEIFS(Datos_Procesados!$E:$E, Datos_Procesados!$B:$B, "<>Sábado", Datos_Procesados!$B:$B, "<>Domingo", Datos_Procesados!$C:$C, Dashboard!$C$3))')
        ws_calc.cell(row=52, column=1, value="Promedio Sab-Dom")
        ws_calc.cell(row=52, column=2, value='=IF(Dashboard!$C$3="Todos", (SUMIF(Datos_Procesados!$B:$B, "Sábado", Datos_Procesados!$E:$E)+SUMIF(Datos_Procesados!$B:$B, "Domingo", Datos_Procesados!$E:$E))/2, (SUMIFS(Datos_Procesados!$E:$E, Datos_Procesados!$B:$B, "Sábado", Datos_Procesados!$C:$C, Dashboard!$C$3)+SUMIFS(Datos_Procesados!$E:$E, Datos_Procesados!$B:$B, "Domingo", Datos_Procesados!$C:$C, Dashboard!$C$3))/2)')

        # --- 3. CONSTRUCCIÓN DE LOS 8 GRÁFICOS ---
        max_r_cron = len(fechas_unicas) + 1 if fechas_unicas else 2

        g1 = BarChart(); g1.title = "🎯 Productividad: Real vs Meta"; g1.style = 10
        g1.add_data(Reference(ws_calc, min_col=2, min_row=1, max_col=3, max_row=max_r_cron), titles_from_data=True)
        g1.set_categories(Reference(ws_calc, min_col=1, min_row=2, max_row=max_r_cron))
        g1.width = 13; g1.height = 9; ws_dash.add_chart(g1, "B5")

        g2 = PieChart(); g2.title = "⚖️ Distribución de Tiempos"
        g2.add_data(Reference(ws_calc, min_col=2, min_row=20, max_row=25), titles_from_data=True)
        g2.set_categories(Reference(ws_calc, min_col=1, min_row=21, max_row=25))
        g2.width = 13; g2.height = 9; ws_dash.add_chart(g2, "H5")

        g3 = DoughnutChart(); g3.title = "⚔️ Carga A100 vs Otras Tareas"; g3.style = 2
        g3.add_data(Reference(ws_calc, min_col=2, min_row=28, max_row=30), titles_from_data=True)
        g3.set_categories(Reference(ws_calc, min_col=1, min_row=29, max_row=30))
        g3.width = 13; g3.height = 9; ws_dash.add_chart(g3, "B19")

        g4 = BarChart(); g4.type = "bar"; g4.title = "🔍 Desglose Operativo: Tareas A100"; g4.legend = None
        g4.add_data(Reference(ws_calc, min_col=2, min_row=33, max_row=37), titles_from_data=True)
        g4.set_categories(Reference(ws_calc, min_col=1, min_row=34, max_row=37))
        g4.width = 13; g4.height = 9; ws_dash.add_chart(g4, "H19")

        g5 = LineChart(); g5.title = "📅 Ritmo Semanal de Carga"; g5.style = 13; g5.legend = None
        g5.add_data(Reference(ws_calc, min_col=2, min_row=40, max_row=47), titles_from_data=True)
        g5.set_categories(Reference(ws_calc, min_col=1, min_row=41, max_row=47))
        g5.width = 13; g5.height = 9; ws_dash.add_chart(g5, "B33")

        g6 = LineChart(); g6.title = "⚡ Eficiencia: Volumen vs Tiempo"
        g6.add_data(Reference(ws_calc, min_col=2, min_row=1, max_col=2, max_row=max_r_cron), titles_from_data=True)
        g6.add_data(Reference(ws_calc, min_col=4, min_row=1, max_col=4, max_row=max_r_cron), titles_from_data=True)
        g6.set_categories(Reference(ws_calc, min_col=1, min_row=2, max_row=max_r_cron))
        g6.width = 13; g6.height = 9; ws_dash.add_chart(g6, "H33")

        g7 = BarChart(); g7.type = "col"; g7.grouping = "stacked"; g7.overlap = 100; g7.title = "🔋 Ocupación vs Disponibilidad"
        g7.add_data(Reference(ws_calc, min_col=4, min_row=1, max_col=5, max_row=max_r_cron), titles_from_data=True)
        g7.set_categories(Reference(ws_calc, min_col=1, min_row=2, max_row=max_r_cron))
        g7.width = 13; g7.height = 9; ws_dash.add_chart(g7, "B47")

        g8 = BarChart(); g8.title = "🏖️ Semana vs Fin de Semana"; g8.legend = None
        g8.add_data(Reference(ws_calc, min_col=2, min_row=50, max_row=52), titles_from_data=True)
        g8.set_categories(Reference(ws_calc, min_col=1, min_row=51, max_row=52))
        g8.width = 13; g8.height = 9; ws_dash.add_chart(g8, "H47")

        # Ocultar las hojas motoras para entregar un archivo limpio
        ws_dash.sheet_properties.tabColor = "1E293B"
        ws_calc.sheet_state = 'hidden'
        ws_flat.sheet_state = 'hidden'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"reporte_consolidado_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al compilar la exportación unificada: {str(e)}")

@router.post("/import/excel", tags=["Metricas"])
async def importar_historico_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")

    try:
        contents = await file.read()
        hojas_a_leer = ['TM', 'TT', 'TN', 'TC']
        
        try:
            xls_data = pd.read_excel(io.BytesIO(contents), sheet_name=hojas_a_leer, engine='openpyxl')
        except ValueError:
            xls_data = pd.read_excel(io.BytesIO(contents), sheet_name=None, engine='openpyxl')
            xls_data = {k: v for k, v in xls_data.items() if k in hojas_a_leer}

        mapa_turnos = {'TM': 'Mañana', 'TT': 'Tarde', 'TN': 'Noche', 'TC': 'Central'}
        registros_agregados = 0
        dias_omitidos = 0

        def clean_val(val):
            if pd.isna(val): return 0
            if isinstance(val, str):
                val = val.strip()
                if val == '-' or val == '': return 0
                try: return int(float(val))
                except: return 0
            return int(val)

        for nombre_hoja, df_hoja in xls_data.items():
            turno_real = mapa_turnos.get(nombre_hoja, "Desconocido")
            
            # MAGIA 1: Quitar espacios invisibles a los nombres de las columnas
            df_hoja.columns = df_hoja.columns.str.strip()
            
            for i in range(0, len(df_hoja), 2):
                if i + 1 >= len(df_hoja): break 
                
                row_c = df_hoja.iloc[i]     # Cantidades
                row_t = df_hoja.iloc[i+1]   # Tiempos

                # MAGIA 2: Filtro inteligente. Extraemos variables clave para ver si hubo actividad
                td_c = clean_val(row_c.get('Tareas Diarias'))
                mails_c = clean_val(row_c.get('Mails'))
                calls_c = clean_val(row_c.get('Calls'))
                
                # Si las tareas principales están en cero, asumimos que el turno aún no ocurrió o está vacío
                if td_c == 0 and mails_c == 0 and calls_c == 0:
                    dias_omitidos += 1
                    continue

                # Extraer fecha
                fecha_raw = row_c.get('Fecha Efectiva')
                if pd.isna(fecha_raw):
                    continue
                try:
                    fecha_dt = pd.to_datetime(fecha_raw, dayfirst=True).to_pydatetime()
                except:
                    fecha_dt = datetime.utcnow()

                nuevo_registro = ShiftMetrics(
                    id=str(uuid.uuid4()),
                    operator_name="Importación Histórica Excel", # Marca clave para el rollback
                    shift=turno_real,
                    created_at=fecha_dt,
                    
                    # Extracción mapeada
                    sh_aws_c=clean_val(row_c.get('SH AWS')), sh_aws_t=clean_val(row_t.get('SH AWS')),
                    sh_c=clean_val(row_c.get('SH')), sh_t=clean_val(row_t.get('SH')),
                    rh_c=clean_val(row_c.get('RH')), rh_t=clean_val(row_t.get('RH')),
                    inout_c=clean_val(row_c.get('IN/OUT')), inout_t=clean_val(row_t.get('IN/OUT')),
                    inc_c=clean_val(row_c.get('INC')), inc_t=clean_val(row_t.get('INC')),
                    inc_aws_c=clean_val(row_c.get('Alarmas/INC AWS')), inc_aws_t=clean_val(row_t.get('Alarmas/INC AWS')),
                    calls_c=calls_c, calls_t=clean_val(row_t.get('Calls')),
                    mails_c=mails_c, mails_t=clean_val(row_t.get('Mails')),
                    td_c=td_c, td_t=clean_val(row_t.get('Tareas Diarias')),
                    acc_c=clean_val(row_c.get('Solicitud de acceso/Visitas/Escoltar de clientes')), acc_t=clean_val(row_t.get('Solicitud de acceso/Visitas/Escoltar de clientes')),
                    trad_c=clean_val(row_c.get('Traducción entre sectores')), trad_t=clean_val(row_t.get('Traducción entre sectores')),
                    pta_c=clean_val(row_c.get('Asistencia de puerta (llaves/elementos,etc)')), pta_t=clean_val(row_t.get('Asistencia de puerta (llaves/elementos,etc)')),
                    mop_c=clean_val(row_c.get('Seguimiento de mantenimiento/Solicitud de MOP')), mop_t=clean_val(row_t.get('Seguimiento de mantenimiento/Solicitud de MOP')),
                    abm_c=clean_val(row_c.get('ABM de Equipos CMDB/ASSET')), abm_t=clean_val(row_t.get('ABM de Equipos CMDB/ASSET')),
                    eti_c=clean_val(row_c.get('ABM de Etiquetas')), eti_t=clean_val(row_t.get('ABM de Etiquetas')),
                    chi_c=clean_val(row_c.get('Llamadas/Mails AWS Chile')), chi_t=clean_val(row_t.get('Llamadas/Mails AWS Chile')),
                    mia_c=clean_val(row_c.get('Llamadas/Casos MIA')), mia_t=clean_val(row_t.get('Llamadas/Casos MIA')),
                    te_c=clean_val(row_c.get('Tareas Adicionales')), te_t=clean_val(row_t.get('Tareas Adicionales'))
                )
                db.add(nuevo_registro)
                registros_agregados += 1

        db.commit()
        return {"mensaje": f"Se importaron {registros_agregados} turnos. Se omitieron {dias_omitidos} filas vacías."}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# MAGIA 3: Endpoint de Rollback
# ==========================================
@router.delete("/import/rollback", tags=["Metricas"])
def deshacer_importacion(db: Session = Depends(get_db)):
    """Elimina únicamente los registros ingresados mediante importación masiva."""
    try:
        registros_a_borrar = db.query(ShiftMetrics).filter(ShiftMetrics.operator_name == "Importación Histórica Excel").delete()
        db.commit()
        return {"mensaje": f"Se han eliminado {registros_a_borrar} registros importados. La base de datos está limpia."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# RUTAS DE EDICIÓN Y BORRADO MANUAL
# ==========================================

@router.delete("/api/{record_id}", tags=["Metricas"])
def delete_metric_record(record_id: str, db: Session = Depends(get_db)):
    """Elimina un registro histórico específico."""
    record = db.query(ShiftMetrics).filter(ShiftMetrics.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    
    db.delete(record)
    db.commit()
    return {"mensaje": "Registro eliminado con éxito"}

@router.get("/api/{record_id}", tags=["Metricas"])
def get_metric_record(record_id: str, db: Session = Depends(get_db)):
    """Obtiene los datos de un registro específico para cargarlos en el modal de edición."""
    record = db.query(ShiftMetrics).filter(ShiftMetrics.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    return record

@router.put("/api/{record_id}", tags=["Metricas"])
async def update_metric_record(record_id: str, request: Request, db: Session = Depends(get_db)):
    """Actualiza los datos de un registro histórico."""
    record = db.query(ShiftMetrics).filter(ShiftMetrics.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    
    body = await request.json()
    
    # Actualizar dinámicamente solo los campos permitidos que vengan en el JSON
    for key, value in body.items():
        if hasattr(record, key) and key not in ['id', 'created_at']: # Protegemos el ID y la fecha
            try:
                setattr(record, key, int(value))
            except ValueError:
                setattr(record, key, 0)
                
    db.commit()
    return {"mensaje": "Tiempos actualizados correctamente"}